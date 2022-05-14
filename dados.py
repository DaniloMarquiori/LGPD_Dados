from openpyxl import Workbook, load_workbook
from openpyxl import Workbook
import csv
import time

# Coletando data para inserção no nome do Excel
data = time.strftime('%Y_%m_%d_%H_%M_%S')

caminho_salvamento = "C:\\Users\\Danil\\Documents\\Pós\\Topicos Avançados\\LGPD\\dados_executados.xlsx"
base = "C:\\Users\\Danil\\Documents\\Pós\\Topicos Avançados\\LGPD\\DadosCSV.csv"

# Criando arquivo
wb = Workbook()
wb.save(caminho_salvamento)

# Defininindo variavel de controle
contador = 0

# Abrindo arquivo criado anteriormente
wb = load_workbook("C:\\Users\\Danil\\Documents\\Pós\\Topicos Avançados\\LGPD\\dados_executados.xlsx")
planilha = wb.worksheets[0]

def random_cpf(cpf):
    if cpf == 'CPF':
        return cpf
    else:
        return cpf + 'sdas'

def random_nome(nome_completo):
    if nome_completo == 'Nome Completo':
        return nome_completo
    else:
        return nome_completo + 'DASDASDASDASDAS'
        
def random_fone(fone):
    if fone == 'Fone':
        return fone
    else:
        return fone + 'DASDASDASDASDAS'

def random_estado_civil(estado_civil):
    if estado_civil == 'Estado Civil':
        return estado_civil
    else:
        return estado_civil + 'DASDASDASDASDAS'

def random_data_nascimento(data_nascimento):
    if data_nascimento == 'Data de Nascimento':
        return data_nascimento
    else:
        return data_nascimento + 'DASDASDASDASDAS'

def random_cep(cep):
    if cep == 'CEP':
        return cep
    else:
        return cep + 'DASDASDASDASDAS'

def random_endereco_residencial(endereco_residencial):
    if endereco_residencial == 'Endereço Residencial':
        return endereco_residencial
    else:
        return endereco_residencial + 'DASDASDASDASDAS'

def random_bairro(bairro):
    if bairro == 'Bairro':
        return bairro
    else:
        return bairro + 'DASDASDASDASDAS'

def random_cidade(cidade):
    if cidade == 'Cidade':
        return cidade
    else:
        return cidade + 'DASDASDASDASDAS'

def random_estado(estado):
    if estado == 'Estado':
        return estado
    else:
        return estado + 'DASDASDASDASDAS'

def random_altura_em_cm(altura_em_cm):
    if altura_em_cm == 'Altura em CM':
        return altura_em_cm
    else:
        return altura_em_cm + 'DASDASDASDASDAS'

def random_numero_do_cartao(numero_do_cartao):
    if numero_do_cartao == 'Número do Cartão':
        return numero_do_cartao
    else:
        return numero_do_cartao + 'DASDASDASDASDAS'

def random_bandeira(bandeira):
    if bandeira == 'Bandeira':
        return bandeira
    else:
        return bandeira + 'DASDASDASDASDAS'

def random_email(email):
    if email == 'Email':
        return email
    else:
        return email + 'DASDASDASDASDAS'


with open(base, newline='', encoding="utf8") as csv_file:
    reader = csv.reader(csv_file)
    while True:
        contador = contador + 1
        try:
            row = next(reader)

            codigo = row[0]
            cpf = row[1]
            nome_completo = row[2]
            sexo = row[3]
            fone = row[4]
            estado_civil = row[5]
            data_de_nascimento = row[6]
            cep = row[7]
            endereço_residencial = row[8]
            bairro = row[9]
            cidade = row[10]
            estado = row[11]
            altura_em_cm = row[12]
            numero_do_cartao = row[13]
            bandeira = row[14]
            email = row[15]

            # print (codigo)
            # print (cpf)
            # print (nome_completo)
            # print (sexo)
            # print (fone)
            # print (estado_civil)
            # print (data_de_nascimento)
            # print (cep)
            # print (endereço_residencial)
            # print (bairro)
            # print (cidade)
            # print (estado)
            # print (altura_em_cm)
            # print (numero_do_cartao)
            # print (bandeira)
            # print (email)

            # Escrever Codigo
            planilha['A'+ (str(contador))] = (str(codigo))

            # Chamar função anonima CPF
            cpf_tratado = random_cpf(cpf)
            planilha['B'+ (str(contador))] = cpf_tratado

            # Chamar função anonima Nome Completo
            nome_tratado = random_nome(nome_completo)
            planilha['C'+ (str(contador))] = nome_tratado

            # Escrever Sexo
            planilha['D'+ (str(contador))] = (str(sexo))

            # Chamar função anonima Fone
            fone_tratado = random_fone(fone)
            planilha['E'+ (str(contador))] = fone_tratado

            # Chamar função anonima Estado Civil
            estado_civil_tratado = random_estado_civil(estado_civil)
            planilha['F'+ (str(contador))] = estado_civil_tratado

            # Chamar função anonima Data de Nascimento
            data_de_nascimento_tratado = random_data_nascimento(data_de_nascimento)
            planilha['G'+ (str(contador))] = data_de_nascimento_tratado

            # Chamar função anonima CEP
            cep_tratado = random_cep(cep)
            planilha['H'+ (str(contador))] = cep_tratado

            # Chamar função anonima Endereço Residencial
            endereço_residencial_tratado = random_endereco_residencial(endereço_residencial)
            planilha['I'+ (str(contador))] = endereço_residencial_tratado

            # Chamar função anonima Bairro
            bairro_tratado = random_bairro(bairro)
            planilha['J'+ (str(contador))] = bairro_tratado

            # Chamar função anonima Cidade
            cidade_tratado = random_cidade(cidade)
            planilha['K'+ (str(contador))] = cidade_tratado

            # Chamar função anonima Estado
            estado_tratado = random_estado(estado)
            planilha['L'+ (str(contador))] = estado_tratado

            # Chamar função anonima Altura em CM
            altura_em_cm_tratado = random_altura_em_cm(altura_em_cm)
            planilha['M'+ (str(contador))] = altura_em_cm_tratado

            # Chamar função anonima Número do Cartão
            numero_do_cartao_tratado = random_numero_do_cartao(numero_do_cartao)
            planilha['N'+ (str(contador))] = numero_do_cartao_tratado

            # Chamar função anonima Bandeira
            bandeira_tratado = random_bandeira(bandeira)
            planilha['O'+ (str(contador))] = bandeira_tratado

            # Chamar função anonima Email
            email_tratado = random_email(email)
            planilha['P'+ (str(contador))] = email_tratado

            # Salvando arquivo
            wb.save("C:\\Users\\Danil\\Documents\\Pós\\Topicos Avançados\\LGPD\\dados_executados.xlsx")

            print ("---------------------------------------------------")

        except csv.Error:
            continue
        except StopIteration:
            break
